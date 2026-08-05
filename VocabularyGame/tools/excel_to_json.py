#!/usr/bin/env python3
"""
excel_to_json.py  ·  CVN Word World
====================================
Converts the CVN vocabulary spreadsheet into data/vocabulary.json.

Sheet layout expected (one sheet per month, e.g. JUNE, JULY, AUG, SEP):

    row 1 :  <blank> | K1 | K2 | K3 | P1 |        |         | P2 | P3 | P4 | P5 | P6
    row 2 :          |    |    |    | Go Get Maths | Science | English
    row 3+:    1     | hello | pen | mom | numbers | plant | pen | camp | ...

Usage
-----
    # every sheet that has words, merged into one file
    python tools/excel_to_json.py Summary_Vocabulary.xlsx data/vocabulary.json

    # a single month only
    python tools/excel_to_json.py Summary_Vocabulary.xlsx data/vocabulary.json --month JUNE

    # one file per month, written next to the output path
    python tools/excel_to_json.py Summary_Vocabulary.xlsx data/vocabulary.json --split
"""

import os
import sys
import json
import argparse
import datetime
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from emoji_map import emoji_for   # noqa: E402  คลัง emoji สำรอง (tools/emoji_map.py)

LEVELS = ["K1", "K2", "K3", "P1", "P2", "P3", "P4", "P5", "P6"]

# ─────────────────────────────────────────────────────────────
#  Emoji fallback — used when assets/<LEVEL>/<word>.png is absent.
#  Mainly matters for K1–K3 where children cannot read yet.
# ─────────────────────────────────────────────────────────────
EMOJI = {
    # people & family
    "hello": "👋", "goodbye": "👋", "teacher": "👩‍🏫", "boy": "👦", "girl": "👧",
    "baby": "👶", "mom": "👩", "dad": "👨", "grandma": "👵", "grandpa": "👴",
    "brother": "👦", "sister": "👧", "family": "👨‍👩‍👧‍👦", "friend": "🧑‍🤝‍🧑",
    "classmate": "🧑‍🤝‍🧑", "guest": "🙋", "doctor": "👨‍⚕️", "judge": "👨‍⚖️",
    "he": "👦", "she": "👧", "who": "❓", "my": "🙋",
    # greetings & school life
    "welcome": "🤗", "thank you": "🙏", "please": "🙏", "polite": "🙇", "greeting": "👋",
    "yes": "✅", "no": "❌", "fine": "👌", "great job": "👏", "help": "🤝",
    "share": "🤝", "wait": "⏳", "line up": "🚶", "stop": "🛑", "first": "1️⃣",
    "school": "🏫", "classroom": "🏫", "desk": "🪑", "chair": "🪑", "table": "🪑",
    "board": "📋", "project": "📊", "slideshow": "📽️", "poster": "📜",
    "flashcard": "🃏", "sticker": "🌟", "phonic": "🔤", "alphabet": "🔤",
    "letter d": "🅳", "letter e": "🅴", "letter f": "🅵", "letter h": "🅷", "letter i": "🅸",
    "name": "🏷️", "sound": "🔊", "music": "🎵", "picture": "🖼️", "art room": "🎨",
    "computer room": "💻", "schoolyard": "🏫", "party hat": "🎉",
    # colours
    "color": "🎨", "colour": "🎨", "blue": "🔵", "yellow": "🟡", "red": "🔴",
    "green": "🟢", "pink": "🩷", "purple": "🟣", "orange": "🟠", "black": "⚫",
    "white": "⚪", "gray": "🩶", "grey": "🩶", "brown": "🟤", "silver": "🥈",
    "shiny": "✨", "rainbow": "🌈", "glitter": "✨",
    # stationery
    "pen": "🖊️", "pencil": "✏️", "crayon": "🖍️", "eraser": "🧽", "paper": "📄",
    "brush": "🖌️", "paint": "🎨", "book": "📖", "bag": "🎒", "ruler": "📏",
    "envelope": "✉️", "newspaper": "📰", "card": "💌", "invitation": "💌",
    "ribbon": "🎀", "wrapping paper": "🎁", "bow": "🎀", "gift bag": "🛍️",
    "confetti": "🎊", "streamer": "🎊", "quilt": "🛏️", "stamp": "📮",
    # animals
    "cat": "🐱", "dog": "🐶", "bird": "🐦", "fish": "🐟", "duck": "🦆",
    "bee": "🐝", "elephant": "🐘", "gorilla": "🦍", "goat": "🐐", "monkey bar": "🐒",
    "spider": "🕷️", "rabbit": "🐰", "frog": "🐸", "frogs": "🐸", "kitten": "🐱",
    "dolphin": "🐬", "penguin": "🐧", "tiger": "🐯", "fox": "🦊", "insect": "🐛",
    "butterfly": "🦋", "feather": "🪶", "teddy bear": "🧸", "zebra": "🦓",
    "giraffe": "🦒", "crocodile": "🐊", "snake": "🐍", "camel": "🐫", "camels": "🐫",
    # food & drink
    "apple": "🍎", "banana": "🍌", "bread": "🍞", "jam": "🍓", "honey": "🍯",
    "juice": "🧃", "milkshake": "🥤", "water": "💧", "cake": "🍰",
    "birthday cake": "🎂", "cookie": "🍪", "candy": "🍬", "chocolate": "🍫",
    "cupcake": "🧁", "ice cream": "🍦", "pizza": "🍕", "sandwich": "🥪",
    "fruit": "🍓", "food": "🍽️", "drink": "🥤", "egg": "🥚", "cup": "☕",
    "glass": "🥛", "fork": "🍴", "spoon": "🥄", "feeding bottle": "🍼",
    "candle": "🕯️", "bean": "🫘", "strawberry": "🍓", "tomato": "🍅", "carrot": "🥕",
    # body
    "eye": "👁️", "ear": "👂", "nose": "👃", "mouth": "👄", "head": "🧑",
    "tummy": "🫃", "hand": "✋", "arm": "💪", "arms": "💪", "leg": "🦵",
    "foot": "🦶", "finger": "☝️", "toe": "🦶", "hair": "💇", "face": "😀",
    "body": "🧍", "elbow": "💪",
    # feelings
    "happy": "😊", "sad": "😢", "tired": "😴", "lively": "🤸", "hungry": "😋",
    "thirsty": "🥤", "surprise": "😲", "dizzy": "😵", "quiet": "🤫",
    "excited": "🤩", "grumpy": "😠", "confident": "😎", "sorry": "😔",
    # actions
    "jump": "🦘", "run": "🏃", "walk": "🚶", "fly": "🕊️", "crawl": "🧎",
    "swim": "🏊", "climb": "🧗", "catch": "🤾", "kick": "🦵", "build": "🔨",
    "sing": "🎤", "dance": "💃", "talk": "🗣️", "sit": "🪑", "sit down": "🪑",
    "look": "👀", "listen": "👂", "smell": "👃", "smile": "😊", "clap": "👏",
    "shake": "🤝", "flap": "🪽", "point": "👉", "draw": "✏️", "trace": "✍️",
    "match": "🔗", "find": "🔍", "guess": "🤔", "count": "🔢", "play": "🎮",
    "ride": "🚲", "race": "🏁", "roll": "🎲", "dip": "🥣", "print": "🖨️",
    "wash": "🧼", "open": "🚪", "knock": "✊", "turn on": "💡", "stretch": "🧘",
    "cross": "❌", "spin": "🌀", "skip": "🤸", "slide": "🛝", "swing": "🛝",
    "blow bubble": "🫧", "splash": "💦", "slip": "🤸", "leave": "👋",
    "celebrate": "🎉", "decorate": "🎨", "walk up": "🪜", "put on": "👕",
    # clothes
    "hat": "🎩", "cap": "🧢", "coat": "🧥", "jacket": "🧥", "shoe": "👟",
    "sock": "🧦", "socks": "🧦", "mittens": "🧤", "glove": "🧤", "scarf": "🧣",
    "boots": "🥾", "boot": "🥾", "sweater": "🧶", "shirt": "👕", "skirt": "👗",
    "shorts": "🩳", "pants": "👖", "raincoat": "🧥", "sunglasses": "🕶️",
    "mask": "🎭", "necklace": "📿", "bead": "📿", "passport": "🛂",
    # toys & play
    "ball": "⚽", "doll": "🪆", "blocks": "🧱", "puzzle": "🧩",
    "jigsaw puzzle": "🧩", "kite": "🪁", "balloon": "🎈", "gift": "🎁",
    "present": "🎁", "drum": "🥁", "toys": "🧸", "toy shop": "🧸",
    "trampoline": "🤸", "seesaw": "🛝", "sandbox": "🏖️", "playground": "🛝",
    "game": "🎮", "treasure hunt": "🗺️", "party": "🎉", "celebration": "🎉",
    "birthday": "🎂", "winner": "🏆",
    # home
    "house": "🏠", "door": "🚪", "window": "🪟", "bed": "🛏️", "pillow": "🛌",
    "blanket": "🛏️", "towel": "🧻", "lamp": "💡", "light": "💡", "sofa": "🛋️",
    "sink": "🚰", "floor": "🟫", "ceiling": "⬜", "stairs": "🪜",
    "kitchen": "🍳", "bedroom": "🛏️", "bathroom": "🛁", "living room": "🛋️",
    "yard": "🌳", "fireplace": "🔥", "fence": "🚧", "garden": "🌷",
    "basket": "🧺", "box": "📦", "soap": "🧼", "telephone": "☎️",
    "radio": "📻", "computer": "💻", "watch": "⌚", "clock": "🕐", "fan": "🪭",
    "apartment": "🏢", "building": "🏢", "villa": "🏡", "tent": "⛺",
    # transport
    "car": "🚗", "bike": "🚲", "train": "🚂", "bus": "🚌", "plane": "✈️",
    "boat": "⛵", "truck": "🚚", "lorry": "🚚", "wheel": "🛞", "tire": "🛞",
    "vehicle": "🚗", "transportation": "🚌", "speed": "💨", "track": "🛤️",
    "bus stop": "🚏", "street sign": "🪧", "traffic light": "🚦",
    "sidewalk": "🚶", "bridge": "🌉", "street": "🛣️", "road": "🛣️",
    # places
    "park": "🏞️", "zoo": "🦁", "library": "📚", "supermarket": "🛒",
    "coffee shop": "☕", "bakery": "🥐", "swimming pool": "🏊", "pool": "🏊",
    "pet store": "🐕", "ice cream shop": "🍦", "shop": "🏪", "town": "🏘️",
    "neighborhood": "🏘️", "farm": "🚜", "lake": "🏞️", "beach": "🏖️",
    "water slide": "🛝", "aquarium": "🐠", "city": "🏙️", "rainforest": "🌴",
    # nature & weather
    "tree": "🌳", "flower": "🌸", "leaf": "🍃", "grass": "🌿", "plant": "🌱",
    "stick": "🪵", "log": "🪵", "rock": "🪨", "stone": "🪨", "moon": "🌙",
    "cloud": "☁️", "cloudy": "☁️", "sun": "☀️", "sunny": "☀️", "rain": "🌧️",
    "rainy": "🌧️", "wind": "💨", "windy": "💨", "snow": "❄️", "snowy": "🌨️",
    "storm": "⛈️", "umbrella": "☂️", "sky": "🌤️", "puddle": "💧",
    "temperature": "🌡️", "weather forecast": "📺", "sled": "🛷",
    "scarecrow": "🎃", "smoke": "💨",
    # numbers
    "one": "1️⃣", "two": "2️⃣", "three": "3️⃣", "four": "4️⃣", "five": "5️⃣",
    "six": "6️⃣", "seven": "7️⃣", "eight": "8️⃣", "nine": "9️⃣", "ten": "🔟",
    "eleven": "1️⃣1️⃣", "twelve": "1️⃣2️⃣", "thirteen": "1️⃣3️⃣",
    "fourteen": "1️⃣4️⃣", "fifteen": "1️⃣5️⃣", "sixteen": "1️⃣6️⃣",
    "seventeen": "1️⃣7️⃣", "eighteen": "1️⃣8️⃣", "nineteen": "1️⃣9️⃣",
    "twenty": "2️⃣0️⃣", "twenty one": "2️⃣1️⃣", "twenty two": "2️⃣2️⃣",
    "twenty three": "2️⃣3️⃣", "twenty four": "2️⃣4️⃣", "twenty five": "2️⃣5️⃣",
    "twenty six": "2️⃣6️⃣", "twenty seven": "2️⃣7️⃣", "twenty eight": "2️⃣8️⃣",
    "twenty nine": "2️⃣9️⃣", "see": "👀",
    # shapes & maths
    "circle": "⭕", "square": "🟦", "triangle": "🔺", "rectangle": "▭",
    "star": "⭐", "heart": "❤️", "oval": "🥚", "shape": "🔷", "pattern": "🔶",
    # opposites & descriptors
    "big": "🐘", "small": "🐜", "tall": "📏", "short": "📏", "old": "👴",
    "new": "✨", "soft": "🧸", "hard": "🪨", "clean": "✨", "dirty": "🧹",
    "hot": "🔥", "cold": "🥶", "warm": "🌤️", "wet": "💧", "dry": "🏜️",
    "fast": "💨", "slow": "🐢", "quickly": "💨", "slowly": "🐢",
    "thin": "📏", "thick": "📚", "missing": "❓", "next": "➡️", "team": "👥",
    "in": "📥", "on": "🔛", "under": "⬇️", "between": "↔️", "idea": "💡",
    "experiment": "🧪", "dusty": "🌫️",
    # days & months
    "sunday": "📅", "monday": "📅", "tuesday": "📅", "wednesday": "📅",
    "thursday": "📅", "friday": "📅", "saturday": "📅",
    "january": "🗓️", "february": "🗓️", "march": "🗓️", "april": "🗓️",
    "may": "🗓️", "june": "🗓️", "july": "🗓️", "august": "🗓️",
    "september": "🗓️", "october": "🗓️", "november": "🗓️", "december": "🗓️",
}


def slug(word: str) -> str:
    """apple -> apple ; air steward -> air_steward"""
    return "".join(
        c for c in word.lower().replace(" ", "_").replace("-", "_")
        if c.isalnum() or c == "_"
    )


# ── ค้นหาไฟล์ภาพ "ของจริง" แทนการเดานามสกุล ──────────────────────────
# เดิมโค้ดเขียน .png ตายตัว แต่ไฟล์จริงเป็น .jpeg เกือบทั้งหมด
# ทำให้เบราว์เซอร์ต้องยิง 404 ทิ้งก่อนทุกภาพ และ preload ใช้ไม่ได้เลย
IMG_EXTS_PREFERRED = (".webp", ".jpeg", ".jpg", ".png", ".avif", ".gif")
_ASSET_INDEX = None


def _assets_root(out_path: str) -> str:
    """assets/ อยู่ระดับเดียวกับ data/vocabulary.json"""
    return os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(out_path))), "assets")


def build_asset_index(assets_dir: str) -> dict:
    """{'K1': {'apple': 'assets/K1/apple.jpeg', ...}, ...}"""
    index = {}
    if not os.path.isdir(assets_dir):
        print(f"  !  ไม่พบโฟลเดอร์ {assets_dir} — จะไม่ใส่ image ให้คำใดเลย")
        return index
    for level in sorted(os.listdir(assets_dir)):
        ldir = os.path.join(assets_dir, level)
        if not os.path.isdir(ldir):
            continue
        found = {}
        for fname in os.listdir(ldir):
            base, ext = os.path.splitext(fname)
            ext = ext.lower()
            if ext not in IMG_EXTS_PREFERRED:
                continue
            key = base.lower()
            prev = found.get(key)
            if prev is None or IMG_EXTS_PREFERRED.index(ext) < IMG_EXTS_PREFERRED.index(
                    os.path.splitext(prev)[1].lower()):
                found[key] = f"assets/{level}/{fname}"
        index[level] = found
    return index


def image_for(level: str, word: str):
    """คืน path ของไฟล์ภาพจริง หรือ None ถ้ายังไม่มีไฟล์"""
    if _ASSET_INDEX is None:
        return None
    return _ASSET_INDEX.get(level, {}).get(slug(word).lower())


def looks_like_note(word: str) -> bool:
    """Teacher notes sometimes sit in the word columns. Filter them out.

    Real vocabulary here is at most three words and never punctuated;
    anything sentence-shaped is a note (e.g. 'Do not add s, es at the end.').
    """
    if len(word) > 30:
        return True
    if any(p in word for p in (".", ",", ";", ":", "!", "?")):
        return True
    if len(word.split()) > 3:
        return True
    return False


def read_sheet(ws):
    """Return {level: [ {word, subject|None}, ... ]} for one month sheet."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return {}, []

    header, subhdr = rows[0], rows[1]

    # Map column index -> (level, subject)
    colmap = {}
    current_level = None
    for i, cell in enumerate(header):
        label = str(cell).strip() if cell not in (None, "") else ""
        if label in LEVELS:
            current_level = label
        if current_level is None:
            continue
        sub = subhdr[i] if i < len(subhdr) else None
        sub = str(sub).strip() if sub not in (None, "") else None
        # a column belongs to the current level if it names it, or if it is a
        # sub-column (blank header + a subject label underneath)
        if label in LEVELS or sub:
            colmap[i] = (current_level, sub)

    out = {lv: [] for lv in LEVELS}
    notes = []
    for r in rows[2:]:
        for i, (level, subject) in colmap.items():
            if i >= len(r):
                continue
            v = r[i]
            if v in (None, ""):
                continue
            word = " ".join(str(v).split())   # collapse stray whitespace
            if not word:
                continue
            if looks_like_note(word):
                notes.append(f"{level}: {word}")
                continue
            out[level].append({"word": word, "subject": subject})
    return out, notes


def convert(xlsx_path, out_path, months=None, split=False):
    global _ASSET_INDEX
    assets_dir = _assets_root(out_path)
    _ASSET_INDEX = build_asset_index(assets_dir)
    print("  ภาพที่พบจริงในโฟลเดอร์ assets:",
          {k: len(v) for k, v in sorted(_ASSET_INDEX.items())})

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    sheets = months or wb.sheetnames
    per_month = {}
    for name in sheets:
        if name not in wb.sheetnames:
            print(f"  !  sheet '{name}' not found — skipped")
            continue
        data, notes = read_sheet(wb[name])
        total = sum(len(v) for v in data.values())
        if total == 0:
            print(f"  ·  {name}: no vocabulary — skipped"
                  + (f"  (notes only: {notes[0]})" if notes else ""))
            continue
        per_month[name] = data
        print(f"  ✓  {name}: {total} words"
              + (f"   [{len(notes)} note cell(s) ignored]" if notes else ""))

    if not per_month:
        sys.exit("No sheets with data found.")

    def build(month_slice, month_label):
        result = {}
        counts = {}
        dupes = []
        for lv in LEVELS:
            seen = {}
            items = []
            for mname, data in month_slice.items():
                for entry in data.get(lv, []):
                    w = entry["word"]
                    key = w.lower()
                    if key in seen:
                        # Same word again. Keep one entry (so a quiz never
                        # shows the word twice) but remember every subject and
                        # month it belongs to, so filters stay accurate.
                        prev = seen[key]
                        if entry["subject"] and entry["subject"] not in prev["subjects"]:
                            prev["subjects"].append(entry["subject"])
                        if mname not in prev["months"]:
                            prev["months"].append(mname)
                        dupes.append(f"{lv}/{w}")
                        continue
                    item = {
                        "word": w,
                        "month": mname,
                        "months": [mname],
                    }
                    # ใส่ image เฉพาะเมื่อมีไฟล์จริงเท่านั้น
                    real_img = image_for(lv, w)
                    if real_img:
                        item["image"] = real_img
                    if entry["subject"]:
                        item["subject"] = entry["subject"]      # first seen (legacy field)
                        item["subjects"] = [entry["subject"]]
                    else:
                        item["subjects"] = []
                    em = EMOJI.get(key) or emoji_for(w)
                    if em:
                        item["emoji"] = em
                    seen[key] = item
                    items.append(item)
            result[lv] = items
            counts[lv] = len(items)

        result["_meta"] = {
            "month": month_label,
            "months": list(month_slice.keys()),
            "levels": LEVELS,
            "counts": counts,
            "total": sum(counts.values()),
            "source": xlsx_path.split("/")[-1],
            "generated": datetime.date.today().isoformat(),
        }
        return result, dupes

    if split:
        base = out_path.rsplit(".", 1)[0]
        for mname, data in per_month.items():
            res, _ = build({mname: data}, mname)
            p = f"{base}_{mname.lower()}.json"
            with open(p, "w", encoding="utf-8") as f:
                json.dump(res, f, ensure_ascii=False, indent=1)
            print(f"  →  {p}  ({res['_meta']['total']} words)")

    label = " + ".join(per_month.keys())
    result, dupes = build(per_month, label)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=1)

    print(f"\n  →  {out_path}")
    print(f"     {result['_meta']['total']} words  ·  " +
          "  ".join(f"{k}:{v}" for k, v in result["_meta"]["counts"].items()))
    missing = sum(1 for lv in ("K1", "K2", "K3")
                  for it in result[lv] if "emoji" not in it)
    print(f"     kindergarten words without emoji fallback: {missing}")
    if dupes:
        print(f"     duplicates removed ({len(dupes)}): " + ", ".join(dupes[:12]) +
              (" ..." if len(dupes) > 12 else ""))


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("out")
    ap.add_argument("--month", action="append",
                    help="sheet name to include (repeatable). Default: all non-empty sheets.")
    ap.add_argument("--split", action="store_true",
                    help="also write one JSON per month")
    a = ap.parse_args()
    convert(a.xlsx, a.out, a.month, a.split)
